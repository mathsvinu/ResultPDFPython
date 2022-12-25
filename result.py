import pandas as pd
import openpyxl
from fpdf import FPDF
def grades(averageMarks):
    if averageMarks >= 90:
        return "A+"
    if averageMarks >= 80 and averageMarks < 90:
        return "A"
    if averageMarks >= 70 and averageMarks < 80:
        return "B+"
    if averageMarks >= 60 and averageMarks < 70:
        return "B"
    if averageMarks >= 50 and averageMarks < 60:
        return "C+"
    if averageMarks >= 25 and averageMarks < 50:
        return "C"
    if averageMarks < 25:
        return "D"

# read by default 1st sheet of an excel file
# dataframe1 = pd.read_excel('IX_A_RESULT.xlsx')
# print(dataframe1)
wb = openpyxl.load_workbook('IX_A_RESULT.xlsx')
sh = wb.active
entireMarksList = []
subjectList = ["Bengali","English","Math","PScience","LScience","History","Geography","Computer"]
top = ("Subjects", "First Term", "Second Term", "Annual Written","Project","Annual Total","Grand Total","Average","Grades")
for i in range(2,sh.max_row+1):
    for j in range(4,12):
        cellvalue = sh.cell(row=i,column=j)
        entireMarksList.append(cellvalue.value)
#print(entireMarksList)
nametag = 2
for roll in range(0,int((sh.max_row-1)/4)):
    print("Roll Number",roll+1," Name:",sh.cell(row=nametag,column=2).value)

    bengaliAnnual = 0
    bengaliTotal = 0
    englishAnnual = 0
    englishTotal = 0
    mathAnnual = 0
    mathTotal = 0
    physcnAnnual = 0
    physcnTotal = 0
    lfscnAnnual = 0
    lfscnTotal = 0
    hisAnnual = 0
    hisTotal = 0
    geoAnnual = 0
    geoTotal = 0
    comAnnual = 0
    comTotal = 0

    bindex = 32*roll
    eindex = 32*roll+1
    mindex = 32*roll+2
    pindex = 32*roll+3
    lindex = 32*roll+4
    hindex = 32*roll+5
    gindex = 32*roll+6
    cindex = 32*roll+7

    subjectAnnualTotal = []
    subjectSessionTotal = []
    firstSummative = []
    secondSummative = []
    thirdSummative = []
    project = []

    # EXTRACTING INDIVIDUAL MARKS FROM THE LIST
    bengaliAnnual += entireMarksList[bindex+2*8]+entireMarksList[bindex+3*8]
    bengaliTotal += bengaliAnnual + entireMarksList[bindex]+entireMarksList[bindex+8]
    subjectAnnualTotal.append(bengaliAnnual)
    subjectSessionTotal.append(bengaliTotal)
    firstSummative.append(entireMarksList[bindex])
    secondSummative.append(entireMarksList[bindex+8])
    thirdSummative.append(entireMarksList[bindex+2*8])
    project.append(entireMarksList[bindex+3*8])

    englishAnnual += entireMarksList[eindex+2*8]+entireMarksList[eindex+3*8]
    englishTotal += englishAnnual + entireMarksList[eindex]+entireMarksList[eindex+8]
    subjectAnnualTotal.append(englishAnnual)
    subjectSessionTotal.append(englishTotal)
    firstSummative.append(entireMarksList[eindex])
    secondSummative.append(entireMarksList[eindex+8])
    thirdSummative.append(entireMarksList[eindex+2*8])
    project.append(entireMarksList[eindex+3*8])

    mathAnnual += entireMarksList[mindex+2*8]+entireMarksList[mindex+3*8]
    mathTotal += mathAnnual + entireMarksList[mindex]+entireMarksList[mindex+8]
    subjectAnnualTotal.append(mathAnnual)
    subjectSessionTotal.append(mathTotal)
    firstSummative.append(entireMarksList[mindex])
    secondSummative.append(entireMarksList[mindex+8])
    thirdSummative.append(entireMarksList[mindex+2*8])
    project.append(entireMarksList[mindex+3*8])

    physcnAnnual += entireMarksList[pindex+2*8]+entireMarksList[pindex+3*8]
    physcnTotal += physcnAnnual + entireMarksList[pindex]+entireMarksList[pindex+8]
    subjectAnnualTotal.append(physcnAnnual)
    subjectSessionTotal.append(physcnTotal)
    firstSummative.append(entireMarksList[pindex])
    secondSummative.append(entireMarksList[pindex+8])
    thirdSummative.append(entireMarksList[pindex+2*8])
    project.append(entireMarksList[pindex+3*8])

    lfscnAnnual += entireMarksList[lindex+2*8]+entireMarksList[lindex+3*8]
    lfscnTotal += lfscnAnnual + entireMarksList[lindex]+entireMarksList[lindex+8]
    subjectAnnualTotal.append(lfscnAnnual)
    subjectSessionTotal.append(lfscnTotal)
    firstSummative.append(entireMarksList[lindex])
    secondSummative.append(entireMarksList[lindex+8])
    thirdSummative.append(entireMarksList[lindex+2*8])
    project.append(entireMarksList[lindex+3*8])

    hisAnnual += entireMarksList[hindex+2*8]+entireMarksList[hindex+3*8]
    hisTotal += hisAnnual + entireMarksList[hindex]+entireMarksList[hindex+8]
    subjectAnnualTotal.append(hisAnnual)
    subjectSessionTotal.append(hisTotal)
    firstSummative.append(entireMarksList[hindex])
    secondSummative.append(entireMarksList[hindex+8])
    thirdSummative.append(entireMarksList[hindex+2*8])
    project.append(entireMarksList[hindex+3*8])

    geoAnnual += entireMarksList[gindex+2*8]+entireMarksList[gindex+3*8]
    geoTotal += geoAnnual + entireMarksList[gindex]+entireMarksList[gindex+8]
    subjectAnnualTotal.append(geoAnnual)
    subjectSessionTotal.append(geoTotal)
    firstSummative.append(entireMarksList[gindex])
    secondSummative.append(entireMarksList[gindex+8])
    thirdSummative.append(entireMarksList[gindex+2*8])
    project.append(entireMarksList[gindex+3*8])

    comAnnual += entireMarksList[cindex+2*8]+entireMarksList[cindex+3*8]
    comTotal += comAnnual + entireMarksList[cindex]+entireMarksList[cindex+8]
    subjectAnnualTotal.append(comAnnual)
    subjectSessionTotal.append(comTotal)
    firstSummative.append(entireMarksList[cindex])
    secondSummative.append(entireMarksList[cindex+8])
    thirdSummative.append(entireMarksList[cindex+2*8])
    project.append(entireMarksList[cindex+3*8])

    # CREATING DATA FOR TABLE PRINTING
    rowOne = []
    rowTwo = []
    rowThree = []
    rowFour = []
    rowFive = []
    rowSix = []
    rowSeven =[]
    rowEight =[]
    for i in range(0,8):
        if i == 0:
            rowOne.append(str(subjectList[i]))
            rowOne.append(str(firstSummative[i]))
            rowOne.append(str(secondSummative[i]))
            rowOne.append(str(thirdSummative[i]))
            rowOne.append(str(project[i]))
            rowOne.append(str(subjectAnnualTotal[i]))
            rowOne.append(str(subjectSessionTotal[i]))
            rowOne.append(str(subjectSessionTotal[i]/2))
            rowOne.append(grades(subjectSessionTotal[i]/2))
        if i == 1:
            rowTwo.append(str(subjectList[i]))
            rowTwo.append(str(firstSummative[i]))
            rowTwo.append(str(secondSummative[i]))
            rowTwo.append(str(thirdSummative[i]))
            rowTwo.append(str(project[i]))
            rowTwo.append(str(subjectAnnualTotal[i]))
            rowTwo.append(str(subjectSessionTotal[i]))
            rowTwo.append(str(subjectSessionTotal[i]/2))
            rowTwo.append(grades(subjectSessionTotal[i]/2))
        if i == 2:
            rowThree.append(str(subjectList[i]))
            rowThree.append(str(firstSummative[i]))
            rowThree.append(str(secondSummative[i]))
            rowThree.append(str(thirdSummative[i]))
            rowThree.append(str(project[i]))
            rowThree.append(str(subjectAnnualTotal[i]))
            rowThree.append(str(subjectSessionTotal[i]))
            rowThree.append(str(subjectSessionTotal[i]/2))
            rowThree.append(grades(subjectSessionTotal[i]/2))
        if i == 3:
            rowFour.append(str(subjectList[i]))
            rowFour.append(str(firstSummative[i]))
            rowFour.append(str(secondSummative[i]))
            rowFour.append(str(thirdSummative[i]))
            rowFour.append(str(project[i]))
            rowFour.append(str(subjectAnnualTotal[i]))
            rowFour.append(str(subjectSessionTotal[i]))
            rowFour.append(str(subjectSessionTotal[i]/2))
            rowFour.append(grades(subjectSessionTotal[i]/2))
        if i == 4:
            rowFive.append(str(subjectList[i]))
            rowFive.append(str(firstSummative[i]))
            rowFive.append(str(secondSummative[i]))
            rowFive.append(str(thirdSummative[i]))
            rowFive.append(str(project[i]))
            rowFive.append(str(subjectAnnualTotal[i]))
            rowFive.append(str(subjectSessionTotal[i]))
            rowFive.append(str(subjectSessionTotal[i]/2))
            rowFive.append(grades(subjectSessionTotal[i]/2))
        if i == 5:
            rowSix.append(str(subjectList[i]))
            rowSix.append(str(firstSummative[i]))
            rowSix.append(str(secondSummative[i]))
            rowSix.append(str(thirdSummative[i]))
            rowSix.append(str(project[i]))
            rowSix.append(str(subjectAnnualTotal[i]))
            rowSix.append(str(subjectSessionTotal[i]))
            rowSix.append(str(subjectSessionTotal[i]/2))
            rowSix.append(grades(subjectSessionTotal[i]/2))
        if i == 6:
            rowSeven.append(str(subjectList[i]))
            rowSeven.append(str(firstSummative[i]))
            rowSeven.append(str(secondSummative[i]))
            rowSeven.append(str(thirdSummative[i]))
            rowSeven.append(str(project[i]))
            rowSeven.append(str(subjectAnnualTotal[i]))
            rowSeven.append(str(subjectSessionTotal[i]))
            rowSeven.append(str(subjectSessionTotal[i]/2))
            rowSeven.append(grades(subjectSessionTotal[i]/2))
        if i == 7:
            rowEight.append(str(subjectList[i]))
            rowEight.append(str(firstSummative[i]))
            rowEight.append(str(secondSummative[i]))
            rowEight.append(str(thirdSummative[i]))
            rowEight.append(str(project[i]))
            rowEight.append(str(subjectAnnualTotal[i]))
            rowEight.append(str(subjectSessionTotal[i]))
            rowEight.append(str(subjectSessionTotal[i]/2))
            rowEight.append(grades(subjectSessionTotal[i]/2))
    data = (top,rowOne,rowTwo,rowThree,rowFour,rowFive,rowSix,rowSeven,rowEight)
    # CREATING MARKSHEET
    pdf = FPDF(format='A4')
    pdf.add_page()
    pdf.set_font("helvetica", size = 15)
    pdf.cell(150, 10, txt = "KUMAR ASHUTOSH INSTITUTION (MAIN) BOYS",border=1,align = "C")
    pdf.ln(15)
    pdf.set_font("helvetica", size = 10)
    pdf.cell(150, 2, txt = "NAME: "+str(sh.cell(row=nametag,column=2).value) + "                     Roll No." + str(roll+1),align = "C")
    
    pdf.set_font_size(8)
    pdf.write_html(
        f"""<table border="1"><thead><tr>
        <th width="9%">{data[0][0]}</th>
        <th width="9%">{data[0][1]}</th>
        <th width="10%">{data[0][2]}</th>
        <th width="12%">{data[0][3]}</th>
        <th width="8%">{data[0][4]}</th>
        <th width="10%">{data[0][5]}</th>
        <th width="10%">{data[0][6]}</th>
        <th width="9%">{data[0][7]}</th>
        <th width="7%">{data[0][8]}</th>
        </tr></thead><tbody><tr>
        <td>{'</td><td>'.join(data[1])}</td>
        </tr><tr>
        <td>{'</td><td>'.join(data[2])}</td>
        </tr><tr>
        <td>{'</td><td>'.join(data[3])}</td>
        </tr><tr>
        <td>{'</td><td>'.join(data[4])}</td>
        </tr><tr>
        <td>{'</td><td>'.join(data[5])}</td>
        </tr><tr>
        <td>{'</td><td>'.join(data[6])}</td>
        </tr><tr>
        <td>{'</td><td>'.join(data[7])}</td>
        </tr><tr>
        <td>{'</td><td>'.join(data[8])}</td>
        </tr></tbody></table>""",
        table_line_separators=True,
    )
    """
    for j in range(0,8):
        stringRow = ""
        print(subjectList[j],firstSummative[j],secondSummative[j],thirdSummative[j],project[j],subjectAnnualTotal[j],subjectSessionTotal[j],subjectSessionTotal[j]/2)
        if len(subjectList[j]) < 22:
            for k in range(0, 22-len(subjectList[j])):
                subjectList[j] += " "
        average = subjectSessionTotal[j]/2
        if average < 25.0:
            stringRow += subjectList[j]+" | "+str(firstSummative[j])+" | "+str(secondSummative[j])+" | "+str(thirdSummative[j])+" | "+str(project[j])+" | "+str(subjectAnnualTotal[j])+" | "+str(subjectSessionTotal[j]) + " | "+str(average) + " | "+"D"
        else:
            stringRow += subjectList[j]+" | "+str(firstSummative[j])+" | "+str(secondSummative[j])+" | "+str(thirdSummative[j])+" | "+str(project[j])+" | "+str(subjectAnnualTotal[j])+" | "+str(subjectSessionTotal[j]) + " | "+str(average)
        pdf.cell(200,10, txt = stringRow, ln = 2, align = 'L')
        pdf.cell(200, 10, txt = "-----------------------------------------------------------------------------",ln = 2, align = 'C')
    #end of j-for loop    """
    pdf.output("Roll No.{0}.pdf".format(roll+1))
    nametag +=4
    print() # END OF ROLL - LOOP
